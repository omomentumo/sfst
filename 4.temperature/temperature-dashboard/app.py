from __future__ import annotations

import csv
import hmac
import io
import json
import os
import re
import sqlite3
import ssl
import threading
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

from flask import (
    Flask,
    Response,
    jsonify,
    redirect,
    request,
    send_file,
    send_from_directory,
    session,
)
from openpyxl import load_workbook
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = BASE_DIR / "frontend"
DB_PATH = BASE_DIR / "temps.db"
REQUIREMENTS_PATH = BASE_DIR / "temperature_requirements.xlsx"

# ==========================================
# Plant-layout monitor (ported from 01MyMonitor / app_gui.py)
# ==========================================
FACTORY_DB_PATH = BASE_DIR / "factory_monitor.db"

# Layout overrides (positions + room-area polygons) edited from the Settings page.
ROOM_CONFIG_PATH = BASE_DIR / "room_config.json"

# Settings page password gate (change via env SETTINGS_PASSWORD).
SETTINGS_PASSWORD = os.environ.get("SETTINGS_PASSWORD", "sats1234")
APP_SECRET_KEY = os.environ.get("APP_SECRET_KEY", "sats-temperature-monitor-secret")
STATS_WINDOW_HOURS = 24

# Live data source (set a valid PHPSESSID via env COOKIE_SESSION when running locally)
COOKIE_SESSION = os.environ.get("COOKIE_SESSION", "PHPSESSID=utbhr9km9kpo3kpmgc1sc5f5be")
WEBSOCKET_URL = os.environ.get("WEBSOCKET_URL", "wss://globallink.itc-group.co.th:9443/ws/")
SITE_ID = os.environ.get("SITE_ID", "3C6AD26B24ECSM")
# Set ENABLE_WS=0 to disable the live WebSocket and only serve stored values.
ENABLE_WS = os.environ.get("ENABLE_WS", "1") not in ("0", "false", "False", "")

# Room map coordinates + per-room max temperature, in the plant-layout image space.
ROOM_LIST = [
    {"room": "1", "x": 653, "y": 479, "max_temp": 25.0},
    {"room": "2", "x": 653, "y": 596, "max_temp": 25.0},
    {"room": "3", "x": 709, "y": 409, "max_temp": 25.0},
    {"room": "4", "x": 709, "y": 453, "max_temp": 25.0},
    {"room": "5", "x": 709, "y": 495, "max_temp": 25.0},
    {"room": "6", "x": 709, "y": 561, "max_temp": 25.0},
    {"room": "7", "x": 709, "y": 602, "max_temp": 25.0},
    {"room": "8", "x": 223, "y": 756, "max_temp": 25.0},
    {"room": "9", "x": 223, "y": 799, "max_temp": 25.0},
    {"room": "10", "x": 223, "y": 713, "max_temp": 25.0},
    {"room": "11", "x": 844, "y": 842, "max_temp": 25.0},
    {"room": "12", "x": 395, "y": 339, "max_temp": 25.0},
    {"room": "13", "x": 458, "y": 339, "max_temp": 25.0},
    {"room": "14", "x": 523, "y": 366, "max_temp": 25.0},
    {"room": "15", "x": 487, "y": 434, "max_temp": 25.0},
    {"room": "16", "x": 487, "y": 568, "max_temp": 25.0},
    {"room": "17", "x": 760, "y": 528, "max_temp": 25.0},
    {"room": "18", "x": 844, "y": 713, "max_temp": 25.0},
    {"room": "19", "x": 604, "y": 348, "max_temp": 25.0},
    {"room": "20", "x": 685, "y": 354, "max_temp": 25.0},
    {"room": "21", "x": 844, "y": 799, "max_temp": 25.0},
    {"room": "22", "x": 844, "y": 756, "max_temp": 25.0},
    {"room": "23", "x": 64, "y": 392, "max_temp": 25.0},
    {"room": "24", "x": 64, "y": 342, "max_temp": 25.0},
    {"room": "25", "x": 64, "y": 597, "max_temp": 25.0},
    {"room": "26", "x": 64, "y": 529, "max_temp": 25.0},
    {"room": "27", "x": 53, "y": 283, "max_temp": 25.0},
    {"room": "28", "x": 53, "y": 465, "max_temp": 25.0},
    {"room": "29", "x": 161, "y": 597, "max_temp": 25.0},
    {"room": "30", "x": 161, "y": 466, "max_temp": 25.0},
    {"room": "31", "x": 85, "y": 231, "max_temp": 25.0},
    {"room": "32", "x": 273, "y": 231, "max_temp": 25.0},
    {"room": "33", "x": 272, "y": 288, "max_temp": 25.0},
    {"room": "34", "x": 230, "y": 322, "max_temp": 25.0},
    {"room": "35", "x": 964, "y": 438, "max_temp": 25.0},
    {"room": "36", "x": 964, "y": 424, "max_temp": 25.0},
    {"room": "37", "x": 964, "y": 495, "max_temp": 10.0},
    {"room": "38", "x": 964, "y": 481, "max_temp": 25.0},
    {"room": "39", "x": 1138, "y": 447, "max_temp": 25.0},
    {"room": "40", "x": 1072, "y": 447, "max_temp": 25.0},
    {"room": "41", "x": 964, "y": 587, "max_temp": 25.0},
    {"room": "42", "x": 964, "y": 540, "max_temp": 25.0},
    {"room": "43", "x": 964, "y": 633, "max_temp": 25.0},
    {"room": "44", "x": 964, "y": 680, "max_temp": 25.0},
    {"room": "45", "x": 1303, "y": 311, "max_temp": 25.0},
    {"room": "46", "x": 1085, "y": 501, "max_temp": 25.0},
    {"room": "47", "x": 1085, "y": 558, "max_temp": 25.0},
    {"room": "48", "x": 1092, "y": 621, "max_temp": 25.0},
    {"room": "49", "x": 757, "y": 373, "max_temp": 25.0},
    {"room": "50", "x": 983, "y": 339, "max_temp": 25.0},
    {"room": "51", "x": 972, "y": 311, "max_temp": 25.0},
    {"room": "52", "x": 972, "y": 268, "max_temp": 25.0},
    {"room": "53", "x": 1124, "y": 308, "max_temp": 25.0},
    {"room": "54", "x": 1198, "y": 364, "max_temp": 25.0},
    {"room": "55", "x": 1018, "y": 444, "max_temp": 25.0},
    {"room": "56", "x": 1131, "y": 415, "max_temp": 25.0},
    {"room": "57", "x": 1059, "y": 341, "max_temp": 25.0},
    {"room": "58", "x": 1124, "y": 341, "max_temp": 25.0},
    {"room": "59", "x": 864, "y": 652, "max_temp": 25.0},
    {"room": "60", "x": 864, "y": 585, "max_temp": 25.0},
    {"room": "61", "x": 864, "y": 518, "max_temp": 25.0},
    {"room": "62", "x": 864, "y": 447, "max_temp": 25.0},
    {"room": "63", "x": 864, "y": 376, "max_temp": 25.0},
    {"room": "64", "x": 1401, "y": 478, "max_temp": 25.0},
    {"room": "65", "x": 1401, "y": 507, "max_temp": 25.0},
    {"room": "66", "x": 1401, "y": 558, "max_temp": 25.0},
    {"room": "67", "x": 1401, "y": 591, "max_temp": 25.0},
    {"room": "68", "x": 1463, "y": 597, "max_temp": 25.0},
    {"room": "69", "x": 1463, "y": 475, "max_temp": 25.0},
    {"room": "70", "x": 1307, "y": 539, "max_temp": 25.0},
    {"room": "71", "x": 1225, "y": 539, "max_temp": 25.0},
    {"room": "72", "x": 1307, "y": 573, "max_temp": 25.0},
    {"room": "73", "x": 1225, "y": 573, "max_temp": 25.0},
    {"room": "74", "x": 1307, "y": 606, "max_temp": 25.0},
    {"room": "75", "x": 1225, "y": 606, "max_temp": 25.0},
    {"room": "76", "x": 1525, "y": 541, "max_temp": 25.0},
    {"room": "77", "x": 1587, "y": 506, "max_temp": 25.0},
    {"room": "78", "x": 1252, "y": 503, "max_temp": 25.0},
    {"room": "79", "x": 1829, "y": 538, "max_temp": 25.0},
    {"room": "80", "x": 1398, "y": 443, "max_temp": 25.0},
    {"room": "81", "x": 1467, "y": 443, "max_temp": 25.0},
    {"room": "82", "x": 1829, "y": 470, "max_temp": 25.0},
    {"room": "83", "x": 1342, "y": 456, "max_temp": 25.0},
    {"room": "84", "x": 1252, "y": 468, "max_temp": 25.0},
    {"room": "85", "x": 1829, "y": 429, "max_temp": 25.0},
    {"room": "86", "x": 1502, "y": 396, "max_temp": 25.0},
    {"room": "87", "x": 1878, "y": 445, "max_temp": 25.0},
    {"room": "88", "x": 1679, "y": 472, "max_temp": 25.0},
    {"room": "89", "x": 1770, "y": 453, "max_temp": 25.0},
    {"room": "90", "x": 1770, "y": 562, "max_temp": 25.0},
    {"room": "91", "x": 1256, "y": 354, "max_temp": 25.0},
    {"room": "92", "x": 1829, "y": 354, "max_temp": 25.0},
    {"room": "93", "x": 1221, "y": 439, "max_temp": 25.0},
    {"room": "94", "x": 1256, "y": 422, "max_temp": 25.0},
    {"room": "95", "x": 1290, "y": 439, "max_temp": 25.0},
    {"room": "96", "x": 1710, "y": 396, "max_temp": 25.0},
]

# room_id -> (idofmach, index_in_value)
TEMP_MAP = {
    "1": ("HR01", 1), "2": ("HR01", 3), "3": ("HR01", 5),
    "4": ("HR01", 8), "5": ("HR01", 11), "6": ("HR01", 14),
    "7": ("HR01", 17), "8": ("HR01", 20), "9": ("HR01", 23),
    "10": ("HR01", 26), "11": ("HR01", 29), "12": ("HR01", 32),
    "13": ("HR01", 34), "14": ("HR01", 36), "15": ("HR01", 38),
    "16": ("HR01", 40), "17": ("HR01", 42), "18": ("HR01", 44),
    "19": ("HR01", 46), "20": ("HR01", 48), "21": ("HR01", 50),
    "22": ("HR01", 52),
    "23": ("OW01", 1), "24": ("OW01", 4), "25": ("OW01", 7),
    "26": ("OW01", 10), "27": ("OW01", 13), "28": ("OW01", 16),
    "29": ("OW01", 19), "30": ("OW01", 21), "31": ("OW01", 23),
    "32": ("OW01", 25), "33": ("OW01", 27), "34": ("OW01", 29),
    "35": ("LR01", 1), "36": ("LR01", 3), "37": ("LR01", 5),
    "38": ("LR01", 7), "39": ("LR01", 9), "40": ("LR01", 11),
    "41": ("LR01", 13), "42": ("LR01", 15), "43": ("LR01", 17),
    "44": ("LR01", 19), "45": ("LR01", 21), "46": ("LR01", 23),
    "47": ("LR01", 25), "48": ("LR01", 27), "49": ("LR01", 29),
    "50": ("LR01", 31), "51": ("LR01", 33), "52": ("LR01", 35),
    "53": ("LR01", 37), "54": ("LR01", 39), "55": ("LR01", 41),
    "56": ("LR01", 43), "57": ("LR01", 45), "58": ("LR01", 47),
    "59": ("LR01", 49), "60": ("LR01", 51), "61": ("LR01", 53),
    "62": ("LR01", 55), "63": ("LR01", 57),
    "64": ("IW01", 1), "65": ("IW01", 4), "66": ("IW01", 7),
    "67": ("IW01", 10), "68": ("IW01", 13), "69": ("IW01", 16),
    "70": ("IW01", 19), "71": ("IW01", 22), "72": ("IW01", 25),
    "73": ("IW01", 28), "74": ("IW01", 31), "75": ("IW01", 34),
    "76": ("IW01", 37), "77": ("IW01", 39), "78": ("IW01", 41),
    "79": ("IW01", 43), "80": ("IW01", 45), "81": ("IW01", 48),
    "82": ("IW01", 51), "83": ("IW01", 53), "84": ("IW01", 55),
    "85": ("IW01", 57), "86": ("IW01", 59), "87": ("IW01", 61),
    "88": ("IW01", 63), "89": ("IW01", 65), "90": ("IW01", 67),
    "91": ("IW01", 69), "92": ("IW01", 71), "93": ("IW01", 73),
    "94": ("IW01", 75), "95": ("IW01", 77), "96": ("IW01", 79),
    "97": ("IW01", 81), "98": ("IW01", 83), "99": ("IW01", 85),
}

ROOM_INDEX = {item["room"]: item for item in ROOM_LIST}
LAYOUT_WIDTH = 1921
LAYOUT_HEIGHT = 729

# In-memory latest reading per room: room -> {"temp": float, "updated": iso}
_latest_lock = threading.Lock()
_latest_temps: dict[str, dict] = {}
_last_save_time: dict[str, datetime] = {}
_ws_state = {"connected": False, "last_message": None}


def as_float(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


# ---------- factory_monitor.db (plant layout history) ----------
def init_factory_db():
    conn = sqlite3.connect(FACTORY_DB_PATH)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS temp_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp DATETIME,
            room_num TEXT,
            temperature REAL
        )
        """
    )
    conn.commit()
    conn.close()


def prime_latest_from_db():
    """Seed the live map with the most recent stored value per room."""
    if not FACTORY_DB_PATH.exists():
        return
    try:
        conn = sqlite3.connect(FACTORY_DB_PATH)
        rows = conn.execute(
            """
            SELECT t.room_num, t.temperature, t.timestamp
            FROM temp_logs t
            JOIN (
                SELECT room_num, MAX(timestamp) AS ts
                FROM temp_logs GROUP BY room_num
            ) latest
            ON t.room_num = latest.room_num AND t.timestamp = latest.ts
            """
        ).fetchall()
        conn.close()
    except sqlite3.Error as exc:
        print(f"Factory DB prime error: {exc}")
        return
    with _latest_lock:
        for room_num, temperature, ts in rows:
            _latest_temps[str(room_num)] = {
                "temp": as_float(temperature),
                "updated": ts,
            }


def save_to_db(room_num, temp_value):
    now = datetime.now()
    if room_num not in _last_save_time or (now - _last_save_time[room_num]).total_seconds() >= 60:
        try:
            conn = sqlite3.connect(FACTORY_DB_PATH)
            conn.execute(
                "INSERT INTO temp_logs (timestamp, room_num, temperature) VALUES (?, ?, ?)",
                (now.strftime("%Y-%m-%d %H:%M:%S"), room_num, temp_value),
            )
            conn.commit()
            conn.close()
            _last_save_time[room_num] = now
        except Exception as exc:
            print(f"DB Error: {exc}")


# ---------- live WebSocket ingester ----------
def map_data(values):
    mapped = {}
    for point in values:
        try:
            mapped[point["id"]] = float(point["v"])
        except (ValueError, TypeError, KeyError):
            continue
    return mapped


def on_message(ws, message):
    import json

    try:
        data = json.loads(message)
        if not isinstance(data, dict):
            return
        if data.get("site") != SITE_ID:
            return

        _ws_state["last_message"] = datetime.now(tz=timezone.utc).isoformat()
        idofmach = data.get("idofmach")
        mapped = map_data(data.get("value", []))

        for room_num, (src_mach, src_id) in TEMP_MAP.items():
            if room_num not in ROOM_INDEX:
                continue
            if idofmach != src_mach:
                continue
            temp_value = mapped.get(src_id)
            if temp_value is None:
                continue
            save_to_db(room_num, temp_value)
            with _latest_lock:
                _latest_temps[room_num] = {
                    "temp": temp_value,
                    "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
    except Exception as exc:
        print(f"Error parsing message: {exc}")


def start_websocket():
    try:
        import websocket  # websocket-client
    except ImportError:
        print("websocket-client not installed; live data disabled. Showing stored values only.")
        return

    headers = {"Cookie": COOKIE_SESSION}
    while True:
        try:
            ws = websocket.WebSocketApp(
                WEBSOCKET_URL,
                header=headers,
                on_open=lambda ws: _ws_state.update(connected=True),
                on_message=on_message,
                on_error=lambda ws, e: print(f"WS Error: {e}"),
                on_close=lambda ws, c, m: _ws_state.update(connected=False),
            )
            ws.run_forever(sslopt={"cert_reqs": ssl.CERT_NONE})
        except Exception as exc:
            print(f"WS Exception: {exc}")
        _ws_state["connected"] = False
        print("Reconnecting in 3s...")
        time.sleep(3)


def load_room_config():
    """Return {room_num: {x?, y?, max_temp?, polygon?}} from the editor's saved file."""
    if not ROOM_CONFIG_PATH.exists():
        return {}
    try:
        with open(ROOM_CONFIG_PATH, "r", encoding="utf-8") as handle:
            data = json.load(handle)
        rooms = data.get("rooms", data) if isinstance(data, dict) else {}
        return rooms if isinstance(rooms, dict) else {}
    except Exception as exc:
        print(f"Room config read error: {exc}")
        return {}


def save_room_config(rooms):
    payload = {"rooms": rooms, "saved_at": datetime.now(tz=timezone.utc).isoformat()}
    tmp = ROOM_CONFIG_PATH.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
    tmp.replace(ROOM_CONFIG_PATH)


def polygon_area(points):
    """Shoelace area in image px^2; points = [[x, y], ...]."""
    if not points or len(points) < 3:
        return 0.0
    total = 0.0
    n = len(points)
    for i in range(n):
        x1, y1 = points[i]
        x2, y2 = points[(i + 1) % n]
        total += (x1 * y2) - (x2 * y1)
    return abs(total) / 2.0


def get_effective_rooms():
    """Merge default ROOM_LIST with saved overrides (position, max_temp, polygon)."""
    overrides = load_room_config()
    merged = []
    for item in ROOM_LIST:
        room_num = item["room"]
        ov = overrides.get(room_num, {}) if isinstance(overrides, dict) else {}
        polygon = ov.get("polygon") or []
        merged.append(
            {
                "room": room_num,
                "x": ov.get("x", item["x"]),
                "y": ov.get("y", item["y"]),
                "max_temp": as_float(ov.get("max_temp")) if ov.get("max_temp") is not None else item["max_temp"],
                "polygon": polygon,
                "area_px": round(polygon_area(polygon), 1) if polygon else None,
            }
        )
    return merged


def build_live_rooms():
    rooms = []
    out_count = 0
    ok_count = 0
    with _latest_lock:
        snapshot = {room: dict(info) for room, info in _latest_temps.items()}
    for item in get_effective_rooms():
        room_num = item["room"]
        reading = snapshot.get(room_num)
        temp = reading.get("temp") if reading else None
        updated = reading.get("updated") if reading else None
        max_temp = item["max_temp"]
        if temp is None:
            status = "NO_DATA"
        elif temp > max_temp:
            status = "OUT"
            out_count += 1
        else:
            status = "OK"
            ok_count += 1
        rooms.append(
            {
                "room": room_num,
                "x": item["x"],
                "y": item["y"],
                "max_temp": max_temp,
                "polygon": item["polygon"],
                "area_px": item["area_px"],
                "temp": temp,
                "status": status,
                "updated": updated,
            }
        )
    return rooms, ok_count, out_count


def compute_24h_stats():
    """Return {room_num: {hi, lo, count, first, last}} over the last STATS_WINDOW_HOURS."""
    stats = {}
    if not FACTORY_DB_PATH.exists():
        return stats
    cutoff = (datetime.now() - timedelta(hours=STATS_WINDOW_HOURS)).strftime("%Y-%m-%d %H:%M:%S")
    try:
        conn = sqlite3.connect(FACTORY_DB_PATH)
        rows = conn.execute(
            """
            SELECT room_num,
                   MIN(temperature) AS lo,
                   MAX(temperature) AS hi,
                   COUNT(*) AS cnt,
                   MIN(timestamp) AS first_ts,
                   MAX(timestamp) AS last_ts
            FROM temp_logs
            WHERE timestamp >= ?
            GROUP BY room_num
            """,
            (cutoff,),
        ).fetchall()
        conn.close()
    except sqlite3.Error as exc:
        print(f"24h stats error: {exc}")
        return stats
    for room_num, lo, hi, cnt, first_ts, last_ts in rows:
        stats[str(room_num)] = {
            "lo": as_float(lo),
            "hi": as_float(hi),
            "count": cnt,
            "first": first_ts,
            "last": last_ts,
        }
    return stats


# ==========================================
# temps.db requirements helpers (existing standalone clone)
# ==========================================
def classify_temperature(actual_temp, requirement, tolerance=2.0):
    actual = as_float(actual_temp)
    required = as_float(requirement)
    if actual is None or required is None:
        return "CRITICAL"
    return "CRITICAL" if actual - required >= tolerance else "OK"


def normalize_lookup_key(value):
    if value is None:
        return None
    text = str(value).strip().upper()
    text = re.sub(r"[^A-Z0-9/:-]", "", text)
    return text or None


def normalize_header(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def find_requirement_column(headers):
    aliases = {
        "requirement", "required", "requiredtemp", "requiredtemperature",
        "setpoint", "setpointtemp", "target", "targettemp",
    }
    for index, header in enumerate(headers):
        if normalize_header(header) in aliases:
            return index
    return None


def find_code_columns(headers):
    aliases = {"unitcode", "sourcecode", "baseroom", "roomcode", "room", "roomid"}
    return [index for index, header in enumerate(headers) if normalize_header(header) in aliases]


def read_requirements_workbook(source):
    workbook = load_workbook(source, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}, 0

    headers = [str(value or "").strip() for value in rows[0]]
    requirement_index = find_requirement_column(headers)
    code_indexes = find_code_columns(headers)
    if requirement_index is None:
        raise ValueError("Excel file needs a Requirement, Setpoint, or Target column")
    if not code_indexes:
        raise ValueError("Excel file needs a Unit Code, Source Code, Base Room, or Room Code column")

    requirements = {}
    row_count = 0
    for row in rows[1:]:
        if requirement_index >= len(row):
            continue
        requirement = as_float(row[requirement_index])
        if requirement is None:
            continue
        matched = False
        for index in code_indexes:
            key = normalize_lookup_key(row[index] if index < len(row) else None)
            if key:
                requirements[key] = requirement
                matched = True
        if matched:
            row_count += 1
    return requirements, row_count


def db_last_synced():
    timestamps = []
    for path in (FACTORY_DB_PATH, DB_PATH, REQUIREMENTS_PATH):
        if path.exists():
            timestamps.append(datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat())
    return max(timestamps, default=None)


# ==========================================
# Flask app + routes
# ==========================================
app = Flask(__name__, static_folder=None)


app.secret_key = APP_SECRET_KEY


@app.after_request
def set_cache_headers(response):
    path = getattr(response, "_request_path", None)
    if path and path.endswith((".html", ".js", ".css")):
        response.cache_control.no_store = True
        response.cache_control.max_age = 0
    return response


@app.route("/")
def root():
    return redirect("/Temperature/index.html")


@app.route("/Temperature/<path:file_path>")
def temperature_assets(file_path):
    response = send_from_directory(FRONTEND_DIR / "Temperature", file_path)
    response._request_path = file_path
    return response


@app.route("/shared/<path:file_path>")
def shared_assets(file_path):
    response = send_from_directory(FRONTEND_DIR / "shared", file_path)
    response._request_path = file_path
    return response


@app.route("/static/<path:file_path>")
def static_assets(file_path):
    return send_from_directory(FRONTEND_DIR / "static", file_path)


@app.route("/api/page-sync/<page_key>")
def page_sync(page_key):
    return jsonify({"page": page_key, "last_synced": db_last_synced()})


# ---- Plant-layout live endpoints (the new Temperature display) ----
@app.route("/api/temperature/live")
def temperature_live():
    rooms, ok_count, out_count = build_live_rooms()
    return jsonify(
        {
            "layout": {"width": LAYOUT_WIDTH, "height": LAYOUT_HEIGHT, "image": "/Temperature/assets/plantlayout_r2.png"},
            "rooms": rooms,
            "summary": {"ok": ok_count, "out": out_count, "total": len(rooms)},
            "ws_connected": bool(_ws_state.get("connected")),
            "last_message": _ws_state.get("last_message"),
            "last_synced": db_last_synced(),
        }
    )


@app.route("/api/temperature/stats")
def temperature_stats():
    return jsonify({"window_hours": STATS_WINDOW_HOURS, "rooms": compute_24h_stats()})


# ---- Settings: password gate + layout config persistence ----
def is_authed():
    return bool(session.get("settings_authed"))


@app.route("/api/settings/status")
def settings_status():
    return jsonify({"authed": is_authed()})


@app.route("/api/settings/login", methods=["POST"])
def settings_login():
    payload = request.get_json(silent=True) or {}
    password = str(payload.get("password", ""))
    if hmac.compare_digest(password, SETTINGS_PASSWORD):
        session["settings_authed"] = True
        session.permanent = True
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Incorrect password"}), 401


@app.route("/api/settings/logout", methods=["POST"])
def settings_logout():
    session.pop("settings_authed", None)
    return jsonify({"ok": True})


@app.route("/api/temperature/config", methods=["GET", "POST"])
def temperature_config():
    if request.method == "GET":
        return jsonify(
            {
                "layout": {
                    "width": LAYOUT_WIDTH,
                    "height": LAYOUT_HEIGHT,
                    "image": "/Temperature/assets/plantlayout_r2.png",
                },
                "rooms": get_effective_rooms(),
            }
        )

    # POST = save (requires auth)
    if not is_authed():
        return jsonify({"error": "Not authorized"}), 401

    payload = request.get_json(silent=True) or {}
    incoming = payload.get("rooms", {})
    if not isinstance(incoming, dict):
        return jsonify({"error": "rooms must be an object keyed by room number"}), 400

    valid_rooms = set(ROOM_INDEX.keys())
    cleaned = {}
    for room_num, conf in incoming.items():
        room_num = str(room_num)
        if room_num not in valid_rooms or not isinstance(conf, dict):
            continue
        entry = {}
        x = as_float(conf.get("x"))
        y = as_float(conf.get("y"))
        if x is not None:
            entry["x"] = round(x, 1)
        if y is not None:
            entry["y"] = round(y, 1)
        max_temp = as_float(conf.get("max_temp"))
        if max_temp is not None:
            entry["max_temp"] = max_temp
        polygon = conf.get("polygon")
        if isinstance(polygon, list):
            pts = []
            for point in polygon:
                if isinstance(point, (list, tuple)) and len(point) == 2:
                    px = as_float(point[0])
                    py = as_float(point[1])
                    if px is not None and py is not None:
                        pts.append([round(px, 1), round(py, 1)])
            if pts:
                entry["polygon"] = pts
        if entry:
            cleaned[room_num] = entry

    try:
        save_room_config(cleaned)
    except Exception as exc:
        return jsonify({"error": f"Could not save config: {exc}"}), 500

    return jsonify({"ok": True, "rooms_saved": len(cleaned)})


@app.route("/api/temperature/history.csv")
def temperature_history_csv():
    if not FACTORY_DB_PATH.exists():
        return jsonify({"error": "No history database found"}), 404
    try:
        conn = sqlite3.connect(FACTORY_DB_PATH)
        rows = conn.execute(
            "SELECT timestamp, room_num, temperature FROM temp_logs ORDER BY timestamp DESC"
        ).fetchall()
        conn.close()
    except sqlite3.Error as exc:
        return jsonify({"error": f"Could not read history: {exc}"}), 500

    buffer = io.StringIO()
    buffer.write("\ufeff")  # UTF-8 BOM for Excel
    writer = csv.writer(buffer)
    writer.writerow(["Date-Time", "Room", "Temperature (C)"])
    writer.writerows(rows)
    filename = f"Temperature_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        buffer.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---- Existing standalone-clone endpoints (kept for compatibility) ----
def load_temperature_requirements():
    if not REQUIREMENTS_PATH.exists():
        return {}
    try:
        requirements, _ = read_requirements_workbook(REQUIREMENTS_PATH)
        return requirements
    except Exception as exc:
        print(f"Temperature requirements workbook error: {exc}")
        return {}


@app.route("/api/temperature/requirements", methods=["GET", "POST"])
def temperature_requirements_file():
    if request.method == "GET":
        if not REQUIREMENTS_PATH.exists():
            return jsonify({"exists": False, "error": "Temperature requirements workbook missing"}), 404
        return send_file(
            REQUIREMENTS_PATH,
            as_attachment=True,
            download_name="temperature_requirements.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return jsonify({"error": "Upload an Excel requirements file"}), 400
    filename = secure_filename(uploaded.filename)
    if not filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Requirements file must be .xlsx"}), 400
    try:
        _, row_count = read_requirements_workbook(uploaded)
        uploaded.seek(0)
        uploaded.save(REQUIREMENTS_PATH)
    except Exception as exc:
        return jsonify({"error": f"Could not read requirements workbook: {exc}"}), 400
    return jsonify({"ok": True, "file": "temperature_requirements.xlsx", "rows": row_count})


@app.route("/api/export/report")
def export_report_placeholder():
    return jsonify({"error": "Report export is not included in this standalone temperature clone."}), 501


def bootstrap():
    init_factory_db()
    prime_latest_from_db()
    if ENABLE_WS:
        thread = threading.Thread(target=start_websocket, daemon=True)
        thread.start()


bootstrap()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5001"))
    app.run(host="127.0.0.1", port=port)
